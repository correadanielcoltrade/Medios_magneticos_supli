#!/usr/bin/env python3
"""Crear archivo Excel de prueba con la nueva estructura de balance"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Datos de ejemplo con la nueva estructura
datos = {
    'Código': [
        '1105050501', '1105050501', '1105050501', '1105050501',
        '1105051002', '1105051002',
        '1110010100', '1110010100', '1110010100',
        '1110020100', '1110020100',
        '2365001', '2366001',
    ],
    'Nombre de la cuenta': [
        'CAJA GENERAL', 'CAJA GENERAL', 'CAJA GENERAL', 'CAJA GENERAL',
        'CAJA PAGOS LOGÍSTICOS', 'CAJA PAGOS LOGÍSTICOS',
        'CUENTA TRANSITORIA', 'CUENTA TRANSITORIA', 'CUENTA TRANSITORIA',
        'COBROS PENDIENTES', 'COBROS PENDIENTES',
        'RETENCIÓN EN LA FUENTE', 'RETENCIÓN EN LA FUENTE',
    ],
    'NIT del tercero': [
        '900534936-5', '222222222-7', '800197268-4', '900538486-0',
        '1007702540', '1073715452',
        '860002964-4', '890903938-8', '1020750460',
        '900898056-0', '1018498580',
        '800197268-4', '890903938-8',
    ],
    'Nombre del tercero': [
        'CÁMARAS Y ALARMAS LAMSEG SAS', 'CUANTÍAS MENORES',
        'DIRECCIÓN DE IMPUESTOS Y ADUANAS NACIONALES', 'ESAR SOLUCIONES LOGÍSTICAS S.A.S',
        'CAJAMARCA MORA JUAN MANUEL', 'SOGAMOSO LATORRE JUAN PABLO',
        'BANCO DE BOGOTÁ', 'BANCOLOMBIA SA', 'ORDOÑEZ ARENAS JUAN CAMILO',
        'MEMORIAS MICROS Y PARTES S A S', 'RAMÍREZ NAVARRO JULIAN DAVID',
        'DIRECCIÓN DE IMPUESTOS Y ADUANAS', 'BANCOLOMBIA SA',
    ],
    'Débito': [
        354648834.14, 350861055.38, 74418823.65, 74160394.96,
        420370.00, 420370.00,
        3032272800.31, 2823881780.17, 287000000.00,
        226000000.00, 226000000.00,
        1500000.00, 2000000.00,
    ],
    'Crédito': [
        0, 0, 0, 0,
        420370.00, 420370.00,
        61000000.00, 61000000.00, 61000000.00,
        0, 0,
        1500000.00, 2000000.00,
    ],
}

df = pd.DataFrame(datos)

# Crear archivo Excel con formato
output_file = 'balance_de_sumas_y_saldos_TEST.xlsx'
df.to_excel(output_file, sheet_name='Balance', index=False)

# Mejorar formato con openpyxl
from openpyxl import load_workbook

wb = load_workbook(output_file)
ws = wb.active

# Aplicar estilos
header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Ajustar anchos
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 35
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 18

# Aplicar formato de número a débito y crédito
for row in ws.iter_rows(min_row=2, min_col=5, max_col=6):
    for cell in row:
        cell.number_format = '#,##0.00'

wb.save(output_file)
print(f"Archivo creado: {output_file}")
print(f"Filas: {len(df)}")
print(f"Columnas: {list(df.columns)}")
