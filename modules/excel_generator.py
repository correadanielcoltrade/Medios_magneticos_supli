from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelGenerator:
    """Genera archivos Excel para los reportes de medios magnéticos."""

    COLORS = {
        'header': 'FF1F4E78',
        'subheader': 'FF4472C4',
        'total': 'FFFFE699',
        'alternado': 'FFDEEBF7'
    }

    FUENTES_CONFIG = {
        'titulo': {'size': 16, 'bold': True, 'color': 'FF1F4E78'},
        'header': {'size': 11, 'bold': True, 'color': 'FFFFFFFF'},
        'data': {'size': 10},
        'total': {'size': 11, 'bold': True}
    }

    def __init__(self, codigo_formato, df_datos, empresa='COLTRADE', anio=2025):
        self.codigo_formato = codigo_formato
        self.df_datos = df_datos
        self.empresa = empresa
        self.anio = anio
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = f"Formato {codigo_formato}"

    def _aplicar_estilos_header(self, fila, columnas):
        fill = PatternFill(start_color=self.COLORS['header'], end_color=self.COLORS['header'], fill_type='solid')
        font = Font(
            size=self.FUENTES_CONFIG['header']['size'],
            bold=self.FUENTES_CONFIG['header']['bold'],
            color=self.FUENTES_CONFIG['header']['color']
        )
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for col_idx, col_name in enumerate(columnas, 1):
            cell = self.ws.cell(row=fila, column=col_idx)
            cell.value = col_name
            cell.fill = fill
            cell.font = font
            cell.border = border
            cell.alignment = alignment

    def _aplicar_estilos_datos(self, fila_inicio=4):
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        max_row = self.ws.max_row
        max_col = self.ws.max_column

        columnas_numericas = set()
        for col_idx in range(1, max_col + 1):
            col_name = self.df_datos.columns[col_idx - 1] if col_idx <= len(self.df_datos.columns) else ""
            if any(texto in col_name for texto in ['Débito', 'Crédito', 'Saldo', 'Balance', 'Movimiento']):
                columnas_numericas.add(col_idx)

        for row_idx in range(fila_inicio, max_row + 1):
            es_total = self.ws.cell(row=row_idx, column=1).value == 'TOTAL'

            for col_idx in range(1, max_col + 1):
                cell = self.ws.cell(row=row_idx, column=col_idx)
                cell.border = border

                if col_idx in columnas_numericas:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.alignment = Alignment(horizontal='left')

                if es_total:
                    cell.fill = PatternFill(start_color=self.COLORS['total'], end_color=self.COLORS['total'], fill_type='solid')
                    cell.font = Font(size=11, bold=True)
                elif row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color=self.COLORS['alternado'], end_color=self.COLORS['alternado'], fill_type='solid')

    def _ajustar_anchos_columnas(self):
        for col_idx, col_name in enumerate(self.df_datos.columns, 1):
            max_length = len(str(col_name))
            for i in range(len(self.df_datos)):
                valor = self.df_datos.iloc[i, col_idx - 1]
                max_length = max(max_length, len(str(valor)))
            self.ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)

    def _generar_excel_formato_1001(self):
        columnas = list(self.df_datos.columns)
        self.wb = Workbook(write_only=True)
        self.ws = self.wb.create_sheet(title=f"Formato {self.codigo_formato}")
        self.ws.freeze_panes = 'A2'
        last_col = get_column_letter(len(columnas))
        self.ws.auto_filter.ref = f"A1:{last_col}{len(self.df_datos) + 1}"
        self._aplicar_anchos_formato_1001()

        fill = PatternFill(start_color=self.COLORS['header'], end_color=self.COLORS['header'], fill_type='solid')
        font = Font(
            size=self.FUENTES_CONFIG['header']['size'],
            bold=self.FUENTES_CONFIG['header']['bold'],
            color=self.FUENTES_CONFIG['header']['color']
        )
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        header = []
        for col_name in columnas:
            cell = WriteOnlyCell(self.ws, value=col_name)
            cell.fill = fill
            cell.font = font
            cell.border = border
            cell.alignment = alignment
            header.append(cell)
        self.ws.append(header)

        for row in self.df_datos.itertuples(index=False, name=None):
            self.ws.append(row)

    def _aplicar_anchos_formato_1001(self):
        anchos = {
            'A': 12, 'B': 18, 'C': 34, 'D': 14, 'E': 24, 'F': 18, 'G': 18,
            'H': 18, 'I': 22, 'J': 30, 'K': 28, 'L': 14, 'M': 14, 'N': 24,
            'O': 18, 'P': 18, 'Q': 18, 'R': 18, 'S': 18, 'T': 18, 'U': 18, 'V': 18
        }
        for col_letter, width in anchos.items():
            self.ws.column_dimensions[col_letter].width = width

    def _generar_excel_formato_1005(self):
        columnas = list(self.df_datos.columns)
        self._aplicar_estilos_header(1, columnas)
        self.ws.row_dimensions[1].height = 32
        self.ws.freeze_panes = 'A2'

        for r_idx, row in enumerate(dataframe_to_rows(self.df_datos, index=False, header=False), 2):
            for c_idx, value in enumerate(row, 1):
                cell = self.ws.cell(row=r_idx, column=c_idx)
                cell.value = value
                if c_idx in (12, 13):
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')

        self.ws.auto_filter.ref = self.ws.dimensions
        self._aplicar_anchos_formato_1005()

    def _aplicar_anchos_formato_1005(self):
        anchos = {
            'A': 22, 'B': 34, 'C': 16, 'D': 16, 'E': 24, 'F': 8, 'G': 22,
            'H': 22, 'I': 22, 'J': 24, 'K': 36, 'L': 20, 'M': 28
        }
        for col_letter, width in anchos.items():
            self.ws.column_dimensions[col_letter].width = width

    def _generar_excel_formato_1006(self):
        columnas = list(self.df_datos.columns)
        self._aplicar_estilos_header(1, columnas)
        self.ws.row_dimensions[1].height = 32
        self.ws.freeze_panes = 'A2'

        for r_idx, row in enumerate(dataframe_to_rows(self.df_datos, index=False, header=False), 2):
            for c_idx, value in enumerate(row, 1):
                cell = self.ws.cell(row=r_idx, column=c_idx)
                cell.value = value
                if 12 <= c_idx <= 14:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')

        self.ws.auto_filter.ref = self.ws.dimensions
        self._aplicar_anchos_formato_1006()

    def _aplicar_anchos_formato_1006(self):
        anchos = {
            'A': 22, 'B': 34, 'C': 16, 'D': 16, 'E': 24, 'F': 8, 'G': 22,
            'H': 22, 'I': 22, 'J': 24, 'K': 36, 'L': 22, 'M': 30, 'N': 22
        }
        for col_letter, width in anchos.items():
            self.ws.column_dimensions[col_letter].width = width

    def _generar_excel_formato_1007(self):
        columnas = list(self.df_datos.columns)
        self._aplicar_estilos_header(1, columnas)
        self.ws.row_dimensions[1].height = 32
        self.ws.freeze_panes = 'A2'

        for r_idx, row in enumerate(dataframe_to_rows(self.df_datos, index=False, header=False), 2):
            for c_idx, value in enumerate(row, 1):
                cell = self.ws.cell(row=r_idx, column=c_idx)
                cell.value = value
                if c_idx in (12, 13):
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')

        self.ws.auto_filter.ref = self.ws.dimensions
        self._aplicar_anchos_formato_1007()

    def _aplicar_anchos_formato_1007(self):
        anchos = {
            'A': 12, 'B': 22, 'C': 34, 'D': 16, 'E': 24, 'F': 22, 'G': 22,
            'H': 22, 'I': 24, 'J': 36, 'K': 12, 'L': 24, 'M': 32
        }
        for col_letter, width in anchos.items():
            self.ws.column_dimensions[col_letter].width = width

    def _generar_excel_formato_1008(self):
        columnas = list(self.df_datos.columns)
        self._aplicar_estilos_header(1, columnas)
        self.ws.row_dimensions[1].height = 32
        self.ws.freeze_panes = 'A2'

        for r_idx, row in enumerate(dataframe_to_rows(self.df_datos, index=False, header=False), 2):
            for c_idx, value in enumerate(row, 1):
                cell = self.ws.cell(row=r_idx, column=c_idx)
                cell.value = value
                if c_idx == 16:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')

        self.ws.auto_filter.ref = self.ws.dimensions
        self._aplicar_anchos_formato_1008()

    def _aplicar_anchos_formato_1008(self):
        anchos = {
            'A': 12, 'B': 24, 'C': 34, 'D': 16, 'E': 24, 'F': 8, 'G': 22,
            'H': 22, 'I': 22, 'J': 24, 'K': 36, 'L': 34, 'M': 12, 'N': 14,
            'O': 12, 'P': 24
        }
        for col_letter, width in anchos.items():
            self.ws.column_dimensions[col_letter].width = width

    def _generar_excel_formato_1009(self):
        columnas = list(self.df_datos.columns)
        self._aplicar_estilos_header(1, columnas)
        self.ws.row_dimensions[1].height = 32
        self.ws.freeze_panes = 'A2'

        for r_idx, row in enumerate(dataframe_to_rows(self.df_datos, index=False, header=False), 2):
            for c_idx, value in enumerate(row, 1):
                cell = self.ws.cell(row=r_idx, column=c_idx)
                cell.value = value
                if c_idx == 16:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')

        self.ws.auto_filter.ref = self.ws.dimensions
        self._aplicar_anchos_formato_1009()

    def _aplicar_anchos_formato_1009(self):
        anchos = {
            'A': 12, 'B': 24, 'C': 38, 'D': 16, 'E': 28, 'F': 8, 'G': 22,
            'H': 22, 'I': 22, 'J': 24, 'K': 40, 'L': 34, 'M': 12, 'N': 14,
            'O': 12, 'P': 24
        }
        for col_letter, width in anchos.items():
            self.ws.column_dimensions[col_letter].width = width

    def _generar_excel_formato_1012(self):
        columnas = list(self.df_datos.columns)
        self._aplicar_estilos_header(1, columnas)
        self.ws.row_dimensions[1].height = 32
        self.ws.freeze_panes = 'A2'

        for r_idx, row in enumerate(dataframe_to_rows(self.df_datos, index=False, header=False), 2):
            for c_idx, value in enumerate(row, 1):
                cell = self.ws.cell(row=r_idx, column=c_idx)
                cell.value = value
                if c_idx == 13:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')

        self.ws.auto_filter.ref = self.ws.dimensions
        self._aplicar_anchos_formato_1012()

    def _aplicar_anchos_formato_1012(self):
        anchos = {
            'A': 12, 'B': 22, 'C': 38, 'D': 18, 'E': 24, 'F': 8, 'G': 22,
            'H': 22, 'I': 22, 'J': 24, 'K': 36, 'L': 24, 'M': 22
        }
        for col_letter, width in anchos.items():
            self.ws.column_dimensions[col_letter].width = width

    def _generar_excel_formato_2276(self):
        columnas = list(self.df_datos.columns)
        self._aplicar_estilos_header(1, columnas)
        self.ws.row_dimensions[1].height = 42
        self.ws.freeze_panes = 'A2'

        for r_idx, row in enumerate(dataframe_to_rows(self.df_datos, index=False, header=False), 2):
            for c_idx, value in enumerate(row, 1):
                cell = self.ws.cell(row=r_idx, column=c_idx)
                cell.value = value
                if 13 <= c_idx <= 41:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')

        self.ws.auto_filter.ref = self.ws.dimensions
        self._aplicar_anchos_formato_2276()

    def _aplicar_anchos_formato_2276(self):
        anchos = {
            'A': 22, 'B': 18, 'C': 22, 'D': 22, 'E': 22, 'F': 22, 'G': 24,
            'H': 36, 'I': 34, 'J': 12, 'K': 14, 'L': 12
        }
        for col_idx in range(13, 42):
            anchos[get_column_letter(col_idx)] = 18
        for col_idx in range(42, 47):
            anchos[get_column_letter(col_idx)] = 22
        for col_letter, width in anchos.items():
            self.ws.column_dimensions[col_letter].width = width

    def _generar_excel_generico(self):
        self.ws['A1'] = f"FORMATO {self.codigo_formato} - MEDIOS MAGNÉTICOS"
        self.ws['A1'].font = Font(**self.FUENTES_CONFIG['titulo'])
        self.ws.merge_cells('A1:I1')
        self.ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws.row_dimensions[1].height = 25

        self.ws['A2'] = f"Empresa: {self.empresa} | Período: {self.anio} | Fecha: {datetime.now().strftime('%d/%m/%Y')}"
        self.ws['A2'].font = Font(size=10, italic=True, color='FF666666')
        self.ws.merge_cells('A2:I2')
        self.ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
        self.ws.row_dimensions[2].height = 18

        columnas = list(self.df_datos.columns)
        self._aplicar_estilos_header(3, columnas)
        self.ws.row_dimensions[3].height = 25

        for r_idx, row in enumerate(dataframe_to_rows(self.df_datos, index=False, header=False), 4):
            for c_idx, value in enumerate(row, 1):
                self.ws.cell(row=r_idx, column=c_idx).value = value

        self._aplicar_estilos_datos(fila_inicio=4)
        self._ajustar_anchos_columnas()

    def generar_excel(self):
        if self.codigo_formato == '1001':
            self._generar_excel_formato_1001()
        elif self.codigo_formato == '1005':
            self._generar_excel_formato_1005()
        elif self.codigo_formato == '1006':
            self._generar_excel_formato_1006()
        elif self.codigo_formato == '1007':
            self._generar_excel_formato_1007()
        elif self.codigo_formato == '1008':
            self._generar_excel_formato_1008()
        elif self.codigo_formato == '1009':
            self._generar_excel_formato_1009()
        elif self.codigo_formato == '2276':
            self._generar_excel_formato_2276()
        elif self.codigo_formato == '1012':
            self._generar_excel_formato_1012()
        else:
            self._generar_excel_generico()

        output = BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output

    def obtener_nombre_archivo(self):
        return f"Formato_{self.codigo_formato}_{self.empresa}_{self.anio}.xlsx"
